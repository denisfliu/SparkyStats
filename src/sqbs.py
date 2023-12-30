import os
from typing import List, Dict
from collections import namedtuple

from openpyxl import load_workbook
from openpyxl.utils.cell import (
    coordinate_from_string,
    column_index_from_string,
    get_column_letter,
)

from src.classes import School
from src.util import *


class Sheet:
    """
    For each sheet, this class collects the necessary information and returns
    the appropriate sqbs string. Almost every leaf in https://www.qbwiki.com/wiki/SQBS_data_file
    is a function.
    """

    sheet_config = get_sheet_config()
    num_sheets_processed = 0

    def __init__(
        self,
        sheet,
        schools: List[School],
        schools_dict: Dict[str, int],
        from_workbook: str,
    ):
        self.sheet = sheet
        self.schools: List[School] = schools
        self.schools_dict: Dict[str, int] = schools_dict
        self.from_workbook = from_workbook
        self.sheet_location_str = f"sheet {sheet.title} from {self.from_workbook}"

        # scorekeeper
        loc = Sheet.sheet_config.scorekeeper
        if loc is None:
            self.scorekeeper = "N/A"
        else:
            self.scorekeeper = self.val(loc)
            if self.scorekeeper is None:
                self.scorekeeper = "N/A"

        self.left_team_config = Sheet.sheet_config.left_team
        self.right_team_config = Sheet.sheet_config.right_team

        self.left_school: School
        self.right_school: School
        self.init_left_and_right_schools()

        self.is_overtime: bool = False
        self.init_overtime()

        self.left_powers, self.left_tens, self.left_negs = 0, 0, 0
        self.right_powers, self.right_tens, self.right_negs = 0, 0, 0
        self.most_tuh_by_player = 0
        self.init_points()

    def compile_sqbs_string(self) -> str:
        fancy_print(f"Starting {self.sheet_location_str}")
        sheet_string = "\n".join(
            [
                self.__id(),
                self.__team_number(left=True),
                self.__team_number(right=True),
                self.__score(left=True),
                self.__score(right=True),
                self.__tossups_heard(),
                self.__round_number(),
                self.__bonuses_heard(left=True),
                self.__bonus_points(left=True),
                self.__bonuses_heard(right=True),
                self.__bonus_points(right=True),
                self.__is_overtime(),
                self.__tu_no_bonus_to_lightning(),
                self.__player_info(),
            ]
        )
        self.__check_issues()
        fancy_print(f"Finished {self.sheet_location_str}")
        return sheet_string

    ########
    ###    Helper functions
    ########
    def get_team_config(self, left: bool = False, right: bool = False) -> str:
        assert left ^ right, "exactly one of left or right should be true"
        return self.left_team_config if left else self.right_team_config

    def init_left_and_right_schools(self) -> None:
        left_team_name = self.val(self.left_team_config.name)
        right_team_name = self.val(self.right_team_config.name)

        if (
            left_team_name not in self.schools_dict.keys()
            or right_team_name not in self.schools_dict.keys()
        ):
            raise ValueError(
                f"one of {left_team_name} or {right_team_name} not found in school list. sheet origin: {self.sheet_location_str}"
            )
        self.left_school = self.schools[self.schools_dict[left_team_name]]
        self.right_school = self.schools[self.schools_dict[right_team_name]]

    def init_points(self) -> None:
        CellPosition = namedtuple(
            "CellPosition", ["name", "tuh", "power", "ten", "neg"]
        )

        max_players = Sheet.sheet_config.max_players_per_team
        lp_config = self.left_team_config.first_player
        rp_config = self.right_team_config.first_player

        def move_right(pos: CellPosition) -> CellPosition:
            # example: move from 'A1' to 'B1'
            increment_func = (
                lambda loc: f"{get_column_letter(column_index_from_string(coordinate_from_string(loc)[0]) + 1)}{coordinate_from_string(loc)[1]}"
            )
            return CellPosition(
                increment_func(pos.name),
                increment_func(pos.tuh),
                increment_func(pos.power),
                increment_func(pos.ten),
                increment_func(pos.neg),
            )

        def fix_none_int(val) -> int:
            if val is None:
                return 0
            return val

        l_pos = CellPosition(
            lp_config.name, lp_config.tuh, lp_config.power, lp_config.ten, lp_config.neg
        )
        r_pos = CellPosition(
            rp_config.name, rp_config.tuh, rp_config.power, rp_config.ten, rp_config.neg
        )

        on_left = True
        player_names = set()
        for pos in (l_pos, r_pos):
            sch = self.left_school if on_left else self.right_school
            for _ in range(max_players):
                name = self.val(pos.name)
                tuh = fix_none_int(self.val(pos.tuh))
                power = fix_none_int(self.val(pos.power))
                ten = fix_none_int(self.val(pos.ten))
                neg = fix_none_int(self.val(pos.neg))
                if name is None or name == "":
                    if tuh > 0 or power > 0 or ten > 0 or neg > 0:
                        self.warning(
                            "no name is getting tossups, skipped point collection"
                        )
                    continue

                num_players = len(player_names)
                player_names.add(name)
                if num_players == len(player_names):
                    self.warning("two of the same player")
                if power + ten + neg > tuh:
                    self.warning("power/ten/neg exceeded tuh")

                sch.find_player(name).add_game_stats(power, ten, neg, tuh)

                if on_left:
                    self.left_powers += power
                    self.left_tens += ten
                    self.left_negs += neg
                else:
                    self.right_powers += power
                    self.right_tens += ten
                    self.right_negs += neg

                self.most_tuh_by_player = max(self.most_tuh_by_player, tuh)
                pos = move_right(pos)

            on_left = False

    def init_overtime(self) -> None:
        if Sheet.sheet_config.overtime is not None:
            self.is_overtime

    # each player is temporarily assigned negs, tens, powers, and tuh per sheet. reset that here
    def reset_player_points(self) -> None:
        for school in (self.left_school, self.right_school):
            for player in school.get_players():
                player.reset_temp_stats()

    # read a value from an excel location
    def val(self, loc: str):
        return self.sheet[loc].value

    def warning(self, warn: str, is_minor: bool = False) -> None:
        bwarning(
            origin=self.sheet_location_str,
            msg=f"(SK/Mod: {self.scorekeeper}): {warn}",
            is_minor=is_minor,
        )

    ########
    ###    Stats sheet stuff
    ########
    def __id(self) -> str:
        Sheet.num_sheets_processed += 1
        return str(Sheet.num_sheets_processed - 1)

    def __team_number(self, left: bool = False, right: bool = False) -> str:
        assert left ^ right, "exactly one of left or right should be true"
        if left:
            return str(self.left_school.get_index())
        return str(self.right_school.get_index())

    def __score(self, left: bool = False, right: bool = False) -> str:
        team_config = self.get_team_config(left, right)
        return str(self.val(team_config.total_score))

    # should be brought up in issues if it doesn't align with tournament setting
    def __tossups_heard(self) -> str:
        general_config = get_general_config()
        tu_heard = general_config.tu_per_game
        if self.is_overtime:
            tu_heard += general_config.tu_per_overtime
        if self.most_tuh_by_player != tu_heard:
            self.warning("")
        return str(tu_heard)

    def __round_number(self) -> str:
        round_number = self.val(Sheet.sheet_config.round_number)
        if round_number is None or round_number == "":
            self.warn(f"bad round number, setting to -1")
            return "-1"
        return str(round_number)

    def __bonuses_heard(self, left: bool = False, right: bool = False) -> str:
        assert left ^ right, "exactly one of left or right should be true"
        if left:
            return self.left_powers + self.left_tens
        if right:
            return self.right_powers + self.right_tens

    def __bonus_points(self, left: bool = False, right: bool = False) -> str:
        assert left ^ right, "exactly one of left or right should be true"
        if left:
            bonus_points = (
                int(self.__score(left))
                - 15 * self.left_powers
                - 10 * self.left_tens
                + 5 * self.left_negs
            )

        if right:
            bonus_points = (
                int(self.__score(right))
                - 15 * self.right_powers
                - 10 * self.right_tens
                + 5 * self.right_negs
            )

        if bonus_points % 10 != 0:
            self.warning(
                f"{self.left_school.get_name() if left else self.right_school.get_name()} have bonuses which are not a multiple of 10"
            )
        return str(bonus_points)

    def __is_overtime(self) -> str:
        return "1" if self.is_overtime else "0"

    # lazy
    def __tu_no_bonus_to_lightning(self) -> str:
        left_team_tossup_without_bonus = "0"
        right_team_tossup_without_bonus = "0"
        forfeit = "0"
        left_lightning = "0"
        right_lightning = "0"
        return "\n".join(
            [
                left_team_tossup_without_bonus,
                right_team_tossup_without_bonus,
                forfeit,
                left_lightning,
                right_lightning,
            ]
        )

    def __player_info(self) -> str:
        player_strings = []
        no_player_str = "-1\n0\n0\n0\n0\n0\n0"
        total_tuh = int(self.__tossups_heard())
        for i in range(8):
            if i + 1 <= self.left_school.get_num_players():
                player = self.left_school.get_players()[i]
                power, ten, neg, tuh = player.get_temp_stats()
                player_strings.append(
                    f"{i}\n{tuh / total_tuh}\n{power}\n{ten}\n{neg}\n0\n{power * 15 + 10 * ten - 5 * neg}"
                )
            else:
                player_strings.append(no_player_str)
            if i + 1 <= self.right_school.get_num_players():
                player = self.right_school.get_players()[i]
                power, ten, neg, tuh = player.get_temp_stats()
                player_strings.append(
                    f"{i}\n{tuh / total_tuh}\n{power}\n{ten}\n{neg}\n0\n{power * 15 + 10 * ten - 5 * neg}"
                )
            else:
                player_strings.append(no_player_str)
        self.reset_player_points()
        return "\n".join(player_strings)

    def __check_issues(self) -> None:
        issues_str = Sheet.sheet_config.issues
        issues = issues_str.split("-")
        warning_list = []
        if len(issues) == 1:
            warning_list.append(self.val(issues[0]))
        elif len(issues) == 2:
            coord1, coord2 = coordinate_from_string(issues[0]), coordinate_from_string(
                issues[1]
            )
            if coord1[0] != coord2[0]:
                raise ValueError(
                    f"Issue Range must be in the same column (same letters)"
                )
            for i in range(coord1[1], coord2[1] + 1):
                warning_list.append(self.val(f"{coord1[0]}{i}"))

        elif len(issues) > 2:
            raise ValueError(
                f"Issues in config cannot be {issues_str}. Must have less than 2 '-'"
            )

        for warning in warning_list:
            if warning is not None and warning != "":
                self.warn(f"[ISSUE] | {issues[0]}")


class Matches:
    """
    Compiles the entire sqbs string together.
    Reads xlsx files from the tournament_directory.
    Note: this class expects a 'roster.xlsx' file.
    https://www.qbwiki.com/wiki/SQBS_data_file
    """

    def __init__(
        self, tournament_directory: str, config_path="tournament_settings.yaml"
    ):
        self.general_config = get_general_config(config_path)
        self.files = [
            os.path.join(tournament_directory, file)
            for file in os.listdir(tournament_directory)
            if file.endswith(".xlsx")
        ]
        roster_sheet = os.path.join(tournament_directory, "roster.xlsx")
        if roster_sheet not in self.files:
            roster_sheet = os.path.join(tournament_directory, "Roster.xlsx")

        assert roster_sheet in self.files, "roster.xlsx or Roster.xlsx required"
        self.files.remove(roster_sheet)

        self.schools: List[School] = list()
        self.schools_dict: Dict[str, int]
        self.sheets = None

        self.init_roster(roster_sheet)
        self.init_sheets()

    def compile_sqbs_string(self) -> str:
        return "\n".join(
            [
                self.__number_of_teams(),
                self.__all_team_information(),
                self.__number_of_matches(),
                self.__matches_strings(),
                self.__points_tracking(),
                self.__reports(),
                self.__use_divisions(),
                self.__sort_method(),
                self.__tournament_name(),
                self.__ftp_settings(),
                self.__always_use_slash(),
                self.__file_suffixes(),
                self.__number_of_divisions(),
                # self.__division_names(),
                self.__number_of_teams(),
                self.__team_division_allocation,
                self.__point_values(),
                self.__packet_information(),
                self.__number_of_teams(),
                self.__exhibition_teams(),
            ]
        )

    def init_roster(self, roster_sheet):
        roster_wb = load_workbook(roster_sheet, data_only=True)
        assert len(roster_wb.sheetnames) == 1, "roster sheet should have one sheet"

        ws = roster_wb.active
        for row in ws.iter_rows():
            school = ""
            players = []
            for i, name in enumerate(row):
                val = name.value
                if val is None:
                    continue
                if i == 0:
                    school = name.value
                else:
                    players.append(name.value)
            if school == "":
                bwarning(
                    origin="Roster Parsing",
                    msg=f"skipped players {players} due to empty school name",
                    is_minor=True,
                )
                continue
            self.schools.append(School(name=school, players=players))
        self.schools.sort(key=lambda school: school.get_name())
        self.schools_dict = {}
        for i, school in enumerate(self.schools):
            self.schools_dict[school.get_name()] = i
            school.index = i

    def init_sheets(self):
        workbooks = [(file, load_workbook(file, data_only=True)) for file in self.files]
        self.sheets = [
            Sheet(
                sheet=sheet,
                schools=self.schools,
                schools_dict=self.schools_dict,
                from_workbook=file,
            )
            for (file, wb) in workbooks
            for sheet in wb
            if sheet.title != "Rosters"
            and sheet.title != "duplicate_me"
            and sheet.title != "Scoresheet"
            and sheet.title != "Instructions"
        ]

    def __all_team_information(self) -> str:
        """
        For each team:
            1 plus number of players
            Team Name
            Each player name on a separate line
        """

        def team_string(school: School) -> str:
            school_str = [str(school.get_num_players() + 1), school.get_name()]
            school_str.extend(school.get_player_strings())
            return "\n".join(school_str)

        return "\n".join([team_string(school) for school in self.schools])

    def __number_of_matches(self) -> str:
        return str(len(self.sheets))

    def __matches_strings(self) -> str:
        return "\n".join([sheet.compile_sqbs_string() for sheet in self.sheets])

    def __points_tracking(self) -> str:
        bonus_conversion = "1"
        bonus_conversion_automatic = "1"
        track_power_and_neg = "3"
        track_lightning = "0"
        track_tuh = "1"
        sort_players_by_pts = "0"
        warning_bit_mask = "254"
        return "\n".join(
            [
                bonus_conversion,
                bonus_conversion_automatic,
                track_power_and_neg,
                track_lightning,
                track_tuh,
                sort_players_by_pts,
                warning_bit_mask,
            ]
        )

    def __reports(self) -> str:
        _round = "1"
        team_standings = "1"
        individual_standings = "1"
        scoreboard = "1"
        team_detail = "1"
        individual_detail = "1"
        stat_key = "1"
        custom_stylesheet = "0"
        return "\n".join(
            [
                _round,
                team_standings,
                individual_standings,
                scoreboard,
                team_detail,
                individual_detail,
                stat_key,
                custom_stylesheet,
            ]
        )

    # too lazy
    def __use_divisions(self) -> str:
        return "0"

    # 1: record, ppg; 2: record, h2h, ppg; 3: record, ss; 4: record, ppth; 5: record, h2h, ppth
    def __sort_method(self) -> str:
        return "4"

    def __tournament_name(self) -> str:
        return self.general_config.name

    def __ftp_settings(self) -> str:
        host_address = ""
        user_name = ""
        directory = ""
        base_file_name = ""
        return "\n".join([host_address, user_name, directory, base_file_name])

    # idk what this means
    def __always_use_slash(self) -> str:
        return "1"

    def __file_suffixes(self) -> str:
        _round = "_rounds.html"
        standings = "_standings.html"
        individuals = "_individuals.html"
        games = "_games.html"
        team_detail = "_teamdetail.html"
        player_detail = "_playerdetail.html"
        statkey = "_statkey.html"
        style_sheet = ""
        return "\n".join(
            [
                _round,
                standings,
                individuals,
                games,
                team_detail,
                player_detail,
                statkey,
                style_sheet,
            ]
        )

    # could add this to the config if desired
    def __number_of_divisions(self) -> str:
        return "0"

    def __division_names(self) -> str:
        raise NotImplementedError

    def __number_of_teams(self) -> str:
        return str(len(self.schools))

    def __team_division_allocation(self) -> str:
        return "\n".join((["-1"] * int(self.__number_of_teams())))

    def __point_values(self) -> str:
        type1 = "15"
        type2 = "10"
        type3 = "-5"
        type4 = "0"
        return "\n".join([type1, type2, type3, type4])

    def __packet_information(self) -> str:
        return "0"

    # enter these in manually
    def __exhibition_teams(self) -> str:
        return "\n".join((["0"] * int(self.__number_of_teams())))
